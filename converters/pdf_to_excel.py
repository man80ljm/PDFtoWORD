"""
PDF转Excel — 从PDF中提取表格数据导出为Excel(.xlsx)文件

使用 pdfplumber 进行表格识别，openpyxl 写入 Excel。
支持：
  - 多页PDF，每页表格写入独立Sheet
  - 可选：所有表格合并到一个Sheet
  - 页范围选择
  - 自动列宽调整
  - 空表格页跳过

通过 on_progress 回调报告进度，不直接操作UI。
"""

import io
import logging
import os
import re
from datetime import datetime

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from core.ocr_client import BaiduOCRClient, REQUESTS_AVAILABLE

# 表格提取策略
TABLE_STRATEGIES = {
    "自动检测": {
        "description": "先按表格线检测，若行数明显偏少自动回退为文本对齐",
    },
    "文本对齐": {
        "description": "按文本位置对齐推断表格（适合无边框/线条断裂表格）",
    },
}


class PDFToExcelConverter:
    """PDF转Excel转换器 — 与 UI 完全解耦。

    用法::

        converter = PDFToExcelConverter(on_progress=my_callback)
        result = converter.convert("input.pdf", strategy="自动检测")
    """

    def __init__(self, on_progress=None):
        self.on_progress = on_progress or (lambda *a: None)

    def _report(self, percent=-1, progress_text="", status_text=""):
        self.on_progress(percent, progress_text, status_text)

    def convert(self, input_file, output_path=None,
                start_page=None, end_page=None,
                strategy="自动检测",
                merge_sheets=False,
                extract_mode="结构提取",
                ocr_mode="平衡",
                api_key=None, secret_key=None):
        """从PDF提取表格并导出为Excel。

        Args:
            input_file: 输入PDF路径
            output_path: 输出xlsx路径，None则自动生成
            start_page: 起始页(1-based)，None=第1页
            end_page: 结束页(1-based)，None=最后一页
            strategy: 表格提取策略（"自动检测" / "文本对齐"）
            merge_sheets: True=所有表格合并到一个Sheet

        Returns:
            dict: success, message, output_file, table_count, total_rows
        """
        result = {
            'success': False,
            'message': '',
            'output_file': '',
            'table_count': 0,
            'total_rows': 0,
        }

        if not PDFPLUMBER_AVAILABLE:
            result['message'] = "pdfplumber 未安装！请执行: pip install pdfplumber"
            return result

        if not OPENPYXL_AVAILABLE:
            result['message'] = "openpyxl 未安装！请执行: pip install openpyxl"
            return result

        if extract_mode == "OCR提取":
            if not REQUESTS_AVAILABLE:
                result['message'] = "requests 未安装，无法使用 OCR 表格识别"
                return result
            if not api_key or not secret_key:
                result['message'] = "已选择 OCR 提取，但未配置百度 OCR API Key/Secret Key"
                return result

        if not input_file or not os.path.exists(input_file):
            result['message'] = "请先选择PDF文件！"
            return result

        if not output_path:
            dir_path = os.path.dirname(input_file)
            basename = os.path.splitext(os.path.basename(input_file))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                dir_path, f"{basename}_表格_{timestamp}.xlsx")

        try:
            self._report(percent=0, progress_text="正在打开PDF...",
                         status_text="读取PDF文件")

            pdf = pdfplumber.open(input_file)
            total_pages = len(pdf.pages)

            if total_pages == 0:
                pdf.close()
                result['message'] = "PDF文件无内容"
                return result

            # 确定页范围（转为0-based索引）
            s_idx = 0
            e_idx = total_pages
            if start_page is not None:
                s_idx = max(0, min(start_page - 1, total_pages - 1))
            if end_page is not None:
                e_idx = max(s_idx + 1, min(end_page, total_pages))

            pages_to_process = e_idx - s_idx

            # 创建 Excel 工作簿
            wb = openpyxl.Workbook()
            # 移除默认Sheet（稍后按需创建）
            wb.remove(wb.active)

            total_tables = 0
            total_rows = 0
            merged_sheet = None
            merged_row_offset = 0

            if merge_sheets:
                merged_sheet = wb.create_sheet(title="所有表格")

            for i, page_idx in enumerate(range(s_idx, e_idx)):
                page = pdf.pages[page_idx]
                page_num = page_idx + 1

                percent = int((i / pages_to_process) * 90)
                self._report(
                    percent=percent,
                    progress_text=f"提取第 {page_num} 页表格... ({percent}%)",
                    status_text=f"第 {page_num}/{total_pages} 页"
                )

                # 提取表格（结构 / OCR）
                if extract_mode == "OCR提取":
                    # 优先尝试结构提取；仅当页面文本弱或结构提取失败时再用 OCR
                    tables = []
                    try:
                        page_text = page.extract_text() or ""
                    except Exception:
                        page_text = ""

                    prefer_structure = self._has_enough_page_text(page_text)
                    if prefer_structure:
                        tables = self._extract_tables(page, strategy)
                    if not tables:
                        tables = self._extract_tables_ocr(
                            page, api_key, secret_key, ocr_mode=ocr_mode
                        )
                else:
                    tables = self._extract_tables(page, strategy)

                if not tables:
                    continue

                for t_idx, table_data in enumerate(tables):
                    if not table_data:
                        continue

                    # 清理表格数据
                    cleaned = self._clean_table(table_data)
                    if not cleaned:
                        continue

                    total_tables += 1

                    if merge_sheets and merged_sheet is not None:
                        # 合并模式：加一行标题行标识来源
                        if merged_row_offset > 0:
                            merged_row_offset += 1  # 表格间空一行

                        # 写入来源标记
                        merged_row_offset += 1
                        cell = merged_sheet.cell(
                            row=merged_row_offset, column=1,
                            value=f"— 第{page_num}页 表格{t_idx + 1} —")
                        cell.font = Font(bold=True, color="4472C4")

                        # 写入表格数据
                        for row in cleaned:
                            merged_row_offset += 1
                            total_rows += 1
                            for col_idx, value in enumerate(row, 1):
                                merged_sheet.cell(
                                    row=merged_row_offset, column=col_idx,
                                    value=value)
                    else:
                        # 独立Sheet模式
                        sheet_name = self._make_sheet_name(
                            f"第{page_num}页", t_idx, wb)
                        ws = wb.create_sheet(title=sheet_name)

                        for row_idx, row in enumerate(cleaned, 1):
                            total_rows += 1
                            for col_idx, value in enumerate(row, 1):
                                ws.cell(row=row_idx, column=col_idx,
                                        value=value)

                        # 样式：首行加粗 + 自动列宽
                        self._style_sheet(ws)

            pdf.close()

            if total_tables == 0:
                result['message'] = (
                    f"在第 {s_idx + 1}~{e_idx} 页中未检测到表格。\n\n"
                    f"建议：\n"
                    f"• 如果PDF是扫描版图片，请先使用「OCR可搜索PDF」功能\n"
                    f"• 尝试切换提取策略为「文本对齐」"
                )
                return result

            # 合并模式下调整样式
            if merge_sheets and merged_sheet is not None:
                self._style_sheet(merged_sheet)

            # 如果没有任何Sheet（不应该发生）
            if len(wb.sheetnames) == 0:
                wb.create_sheet(title="空")

            # 保存
            self._report(percent=92, progress_text="正在保存Excel...",
                         status_text="写入xlsx文件")
            wb.save(output_path)

            result['success'] = True
            result['output_file'] = output_path
            result['table_count'] = total_tables
            result['total_rows'] = total_rows
            result['message'] = (
                f"提取完成！\n"
                f"处理了 {pages_to_process} 页，发现 {total_tables} 个表格\n"
                f"共 {total_rows} 行数据"
            )
            self._report(percent=100, progress_text="提取完成！")

        except Exception as e:
            logging.error(f"PDF转Excel失败: {e}", exc_info=True)
            result['message'] = f"转换失败：{str(e)}"

        return result

    @staticmethod
    def _build_table_settings(strategy):
        """根据策略构建 pdfplumber 表格提取参数"""
        if strategy == "文本对齐":
            return {
                "vertical_strategy": "text",
                "horizontal_strategy": "text",
                "snap_tolerance": 5,
                "join_tolerance": 5,
                "text_tolerance": 2,
                "intersection_tolerance": 5,
                "min_words_vertical": 2,
                "min_words_horizontal": 2,
            }
        else:
            # 自动检测（默认）：优先用线条，回退到文本
            return {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 4,
                "join_tolerance": 4,
                "intersection_tolerance": 5,
            }

    def _extract_tables(self, page, strategy):
        """按策略提取表格，自动检测会在行数偏少时回退到文本对齐。"""
        if strategy == "文本对齐":
            settings = self._build_table_settings("文本对齐")
            return page.extract_tables(settings)

        # 自动检测：先线条
        line_settings = self._build_table_settings("自动检测")
        line_tables = page.extract_tables(line_settings) or []
        line_rows = self._count_rows(line_tables)

        # 再文本对齐
        text_settings = self._build_table_settings("文本对齐")
        text_tables = page.extract_tables(text_settings) or []
        text_rows = self._count_rows(text_tables)

        if text_rows > line_rows:
            return text_tables
        return line_tables

    @staticmethod
    def _count_rows(tables):
        count = 0
        for table in tables:
            if not table:
                continue
            for row in table:
                if row is None:
                    continue
                if any(cell for cell in row):
                    count += 1
        return count

    def _extract_tables_ocr(self, page, api_key, secret_key, ocr_mode="平衡"):
        """使用 OCR 表格识别，返回二维表格列表。"""
        client = BaiduOCRClient(api_key, secret_key)
        resolution = self._ocr_mode_to_resolution(ocr_mode)
        try:
            page_img = page.to_image(resolution=resolution).original
            buf = io.BytesIO()
            page_img.save(buf, format="PNG")
            img_bytes = buf.getvalue()
        except Exception as e:
            raise RuntimeError(f"无法渲染页面用于OCR: {e}")
        result = client.recognize_table(img_bytes, return_excel=False, cell_contents=False)
        tables = []
        tables_result = result.get("tables_result", [])
        for table in tables_result:
            body = table.get("body", [])
            grid = self._table_body_to_grid(body)
            if grid:
                grid = self._normalize_ocr_table(grid)
            if grid:
                tables.append(grid)
        return tables

    @staticmethod
    def _table_body_to_grid(body):
        if not body:
            return []
        max_row = 0
        max_col = 0
        for cell in body:
            max_row = max(max_row, cell.get("row_end", 0))
            max_col = max(max_col, cell.get("col_end", 0))
        if max_row <= 0 or max_col <= 0:
            return []
        grid = [["" for _ in range(max_col)] for _ in range(max_row)]
        for cell in body:
            r = cell.get("row_start", 0) - 1
            c = cell.get("col_start", 0) - 1
            if r < 0 or c < 0:
                continue
            text = cell.get("words", "")
            if grid[r][c]:
                grid[r][c] = f"{grid[r][c]} {text}".strip()
            else:
                grid[r][c] = text
        return grid

    @staticmethod
    def _normalize_ocr_table(grid):
        """对 OCR 表格做表头识别 + 自动对齐。"""
        if not grid:
            return grid

        header_idx, header = PDFToExcelConverter._find_header_row(grid)
        if header_idx is None:
            # 没找到表头，仍做空列裁剪
            return PDFToExcelConverter._trim_empty_columns(grid)

        # 仅保留表头及其后数据
        data = grid[header_idx:]
        data = PDFToExcelConverter._align_rows_to_header(data, len(header))
        data = PDFToExcelConverter._trim_empty_columns(data)
        return data

    @staticmethod
    def _find_header_row(grid):
        keywords = ["班级", "学号", "姓名", "平时", "期中", "期末", "总评", "备注", "成绩"]
        best_idx = None
        best_score = 0
        best_row = None
        scan_rows = min(len(grid), 8)
        for i in range(scan_rows):
            row = grid[i]
            if not row:
                continue
            text = " ".join(str(c) for c in row if c)
            score = 0
            for kw in keywords:
                if kw in text:
                    score += 1
            if score > best_score:
                best_score = score
                best_idx = i
                best_row = row
        if best_score >= 2:
            return best_idx, best_row
        return None, None

    @staticmethod
    def _align_rows_to_header(rows, header_len):
        aligned = []
        for row in rows:
            if not row:
                continue
            # 去掉末尾空单元
            while row and not str(row[-1]).strip():
                row = row[:-1]
            # 自动左移：如果前面空列过多
            leading_empty = 0
            for c in row:
                if str(c).strip():
                    break
                leading_empty += 1
            if leading_empty >= 1 and len(row) - leading_empty <= header_len:
                row = row[leading_empty:]
            # 统一列数
            if len(row) < header_len:
                row = row + [""] * (header_len - len(row))
            elif len(row) > header_len:
                row = row[:header_len]
            aligned.append(row)
        return aligned

    @staticmethod
    def _trim_empty_columns(grid):
        if not grid:
            return grid
        col_count = max(len(r) for r in grid)
        keep = []
        for c in range(col_count):
            has_value = False
            for r in grid:
                if c < len(r) and str(r[c]).strip():
                    has_value = True
                    break
            if has_value:
                keep.append(c)
        trimmed = []
        for r in grid:
            trimmed.append([r[c] if c < len(r) else "" for c in keep])
        return trimmed

    @staticmethod
    def _has_enough_page_text(text, min_chars=24):
        raw = (text or "").strip()
        if not raw:
            return False
        compact = "".join(raw.split())
        if len(compact) < min_chars:
            return False
        effective = sum(1 for ch in compact if ch.isalnum() or ("\u4e00" <= ch <= "\u9fff"))
        return effective >= max(12, min_chars // 2)

    @staticmethod
    def _ocr_mode_to_resolution(ocr_mode):
        mode = (ocr_mode or "平衡").strip()
        mapping = {
            "快速": 220,
            "平衡": 300,
            "高精": 360,
        }
        return mapping.get(mode, 300)

    @staticmethod
    def _clean_table(table_data):
        """清理表格数据：去除空行、规范化单元格内容"""
        cleaned = []
        for row in table_data:
            if row is None:
                continue
            clean_row = []
            for cell in row:
                if cell is None:
                    clean_row.append("")
                else:
                    # 合并换行、去除多余空白
                    text = str(cell).strip()
                    text = ' '.join(text.split())
                    clean_row.append(text)
            # 跳过全空行
            if any(c for c in clean_row):
                cleaned.append(clean_row)
        return cleaned

    @staticmethod
    def _make_sheet_name(base_name, table_idx, wb):
        """生成唯一的Sheet名称（Excel限31字符）"""
        if table_idx == 0:
            name = base_name
        else:
            name = f"{base_name}_表{table_idx + 1}"

        # Excel Sheet名最长31字符
        name = name[:31]

        # 确保唯一
        existing = set(wb.sheetnames)
        if name not in existing:
            return name
        suffix = 2
        while True:
            candidate = f"{name[:28]}_{suffix}"
            if candidate not in existing:
                return candidate
            suffix += 1

    @staticmethod
    def _style_sheet(ws):
        """为工作表添加样式：首行加粗、自动列宽、边框"""
        if ws.max_row is None or ws.max_row == 0:
            return

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

        header_font = Font(bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical='center')

        # 遍历所有单元格
        for row_idx, row in enumerate(ws.iter_rows(
                min_row=1, max_row=ws.max_row,
                max_col=ws.max_column), 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = wrap_alignment
                if row_idx == 1:
                    cell.font = header_font

        # 自动列宽
        for col_idx in range(1, (ws.max_column or 0) + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.value:
                    # 估算中文字符宽度（中文字符约占2个英文字符宽度）
                    text = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in text)
                    max_len = max(max_len, char_len)
            # 限制列宽范围
            adjusted_width = min(max(max_len + 2, 8), 60)
            ws.column_dimensions[col_letter].width = adjusted_width
